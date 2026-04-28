"""
Build a batch of Postgres MCP family-metrics queries for Gate Q2/Q3/Q4 inputs.

Usage:
  python tools/family_metrics_postgres_batch.py "C:\\path\\to\\filled_workbook.xlsx"
  python tools/family_metrics_postgres_batch.py "C:\\path\\to\\output\\research_sessions\\session_name"
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import UTC, date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sku_lookup import strip_pack_suffix
from template_parser import DEFAULT_WORKBOOK, SHEET_NAME, normalize_header


AMAZON_CHANNEL_ID = 11929
SHOPIFY_CHANNEL_ID = 12585


def utc_now() -> str:
    """Return an ISO timestamp for artifact generation."""
    return datetime.now(UTC).replace(microsecond=0).isoformat()


def default_monthly_window(anchor_date: date | None = None) -> tuple[str, str]:
    """Return the last 24 complete calendar months."""
    anchor = anchor_date or date.today()
    end_date = date(anchor.year, anchor.month, 1)
    start_date = date(end_date.year - 2, end_date.month, 1)
    return start_date.isoformat(), end_date.isoformat()


def default_customer_window(anchor_date: date | None = None) -> tuple[str, str]:
    """Return the last 12 complete calendar months."""
    anchor = anchor_date or date.today()
    end_date = date(anchor.year, anchor.month, 1)
    start_date = date(end_date.year - 1, end_date.month, 1)
    return start_date.isoformat(), end_date.isoformat()


def output_stem_for(target: Path) -> Path:
    """Choose a reasonable default output location for generated files."""
    if target.is_dir():
        return target / "family_metrics"
    return target.with_suffix("")


def collect_reference_rows_from_session(session_dir: Path) -> list[dict[str, Any]]:
    """Read unique reference SKUs from session packet files."""
    packets_dir = session_dir / "packets"
    references: dict[str, dict[str, Any]] = {}

    for packet_file in sorted(packets_dir.glob("row_*_packet.json")):
        packet = json.loads(packet_file.read_text(encoding="utf-8"))
        identity = packet.get("identity") or {}
        sku = identity.get("sunco_reference_sku")
        row_number = packet.get("row_number")
        if not sku or row_number is None:
            continue
        normalized_sku = str(sku).strip().upper()
        entry = references.setdefault(
            normalized_sku,
            {
                "sku": normalized_sku,
                "family": strip_pack_suffix(normalized_sku),
                "row_numbers": [],
            },
        )
        if row_number not in entry["row_numbers"]:
            entry["row_numbers"].append(int(row_number))

    return sorted(references.values(), key=lambda item: item["row_numbers"][0])


def collect_reference_rows_from_workbook(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Read unique reference SKUs directly from a workbook."""
    workbook = load_workbook(workbook_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

    ws = workbook[sheet_name]
    header_index: dict[int, str] = {}

    for col in range(1, ws.max_column + 1):
        header_key = normalize_header(ws.cell(row=2, column=col).value)
        if header_key:
            header_index[col] = header_key

    sku_columns = [idx for idx, key in header_index.items() if key == "sunco_reference_sku"]
    if not sku_columns:
        raise ValueError("Could not find a 'Sunco Reference SKU' column in the workbook.")
    sku_column = sku_columns[0]

    references: dict[str, dict[str, Any]] = {}
    for row in range(3, ws.max_row + 1):
        raw_value = ws.cell(row=row, column=sku_column).value
        if raw_value is None:
            continue
        normalized_sku = str(raw_value).strip().upper()
        if not normalized_sku:
            continue
        entry = references.setdefault(
            normalized_sku,
            {
                "sku": normalized_sku,
                "family": strip_pack_suffix(normalized_sku),
                "row_numbers": [],
            },
        )
        entry["row_numbers"].append(row)

    return sorted(references.values(), key=lambda item: item["row_numbers"][0])


def collect_reference_rows(target: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Collect unique reference SKUs from either a session dir or workbook."""
    if target.is_dir():
        packets_dir = target / "packets"
        if packets_dir.exists():
            return collect_reference_rows_from_session(target)
        raise ValueError(f"Directory '{target}' is not a research session root.")
    return collect_reference_rows_from_workbook(target, sheet_name)


def build_monthly_sales_query(family: str, start_date: str, end_date: str) -> str:
    """Build the Postgres MCP monthly sales query."""
    return f"""
SELECT DATE_TRUNC('month', o.order_date) AS month,
       o.sales_channel_id,
       SUM(oi.sales_price * oi.quantity_ordered) AS revenue,
       SUM(oi.quantity_ordered) AS units,
       COUNT(DISTINCT o.customer_id) AS distinct_customers
FROM skubana_order o
JOIN skubana_orderitem oi ON o.order_id = oi.order_id
WHERE oi.listing_sku LIKE '{family}%'
  AND o.order_date >= '{start_date}'
  AND o.order_date < '{end_date}'
  AND o.order_status NOT IN ('CANCELLED')
  AND o.sales_channel_id IN ({SHOPIFY_CHANNEL_ID}, {AMAZON_CHANNEL_ID})
GROUP BY DATE_TRUNC('month', o.order_date), o.sales_channel_id
ORDER BY month, o.sales_channel_id
""".strip()


def build_customer_concentration_query(family: str, start_date: str, end_date: str) -> str:
    """Build the Postgres MCP customer concentration query."""
    return f"""
SELECT o.sales_channel_id,
       o.customer_id,
       SUM(oi.sales_price * oi.quantity_ordered) AS customer_revenue,
       SUM(oi.quantity_ordered) AS customer_units,
       COUNT(DISTINCT DATE_TRUNC('month', o.order_date)) AS order_months
FROM skubana_order o
JOIN skubana_orderitem oi ON o.order_id = oi.order_id
WHERE oi.listing_sku LIKE '{family}%'
  AND o.order_date >= '{start_date}'
  AND o.order_date < '{end_date}'
  AND o.order_status NOT IN ('CANCELLED')
  AND o.sales_channel_id IN ({SHOPIFY_CHANNEL_ID}, {AMAZON_CHANNEL_ID})
GROUP BY o.sales_channel_id, o.customer_id
ORDER BY o.sales_channel_id, customer_revenue DESC
""".strip()


def build_payload_template_row(sku: str, family: str, row_numbers: list[int], monthly_start: str, monthly_end: str, customer_start: str, customer_end: str) -> dict[str, Any]:
    """Return the merge-ready payload skeleton accepted by --family-metrics-json."""
    return {
        "sku": sku,
        "family": family,
        "row_numbers": row_numbers,
        "family_metrics_source": "postgres_mcp",
        "family_metrics_period_label": {
            "monthly_sales_24mo": {"start_date": monthly_start, "end_date": monthly_end},
            "customer_concentration_12mo": {"start_date": customer_start, "end_date": customer_end},
        },
        "monthly_sales": [],
        "customer_concentration": {
            "amazon": {
                "total_customers": None,
                "total_revenue": None,
                "top_20pct_customer_count": None,
                "top_20pct_revenue": None,
                "top_20pct_revenue_share_pct": None,
                "repeat_customer_count": None,
                "repeat_rate_pct": None,
            },
            "shopify": {
                "total_customers": None,
                "total_revenue": None,
                "top_20pct_customer_count": None,
                "top_20pct_revenue": None,
                "top_20pct_revenue_share_pct": None,
                "repeat_customer_count": None,
                "repeat_rate_pct": None,
            },
        },
        "family_metrics_notes": [],
    }


def build_query_bundle(
    target: Path,
    sheet_name: str,
    monthly_start_date: str | None,
    monthly_end_date: str | None,
    customer_start_date: str | None,
    customer_end_date: str | None,
) -> dict[str, Any]:
    """Build the batch query payload plus a merge-ready template."""
    references = collect_reference_rows(target, sheet_name)
    resolved_monthly_start, resolved_monthly_end = (
        (monthly_start_date, monthly_end_date)
        if monthly_start_date and monthly_end_date
        else default_monthly_window()
    )
    resolved_customer_start, resolved_customer_end = (
        (customer_start_date, customer_end_date)
        if customer_start_date and customer_end_date
        else default_customer_window()
    )

    items = []
    payload_template = []
    row_map: dict[int, list[str]] = defaultdict(list)
    for ref in references:
        sku = ref["sku"]
        family = ref["family"]
        row_numbers = sorted(ref["row_numbers"])
        items.append(
            {
                "sku": sku,
                "family": family,
                "row_numbers": row_numbers,
                "queries": {
                    "monthly_sales_24mo": build_monthly_sales_query(
                        family,
                        resolved_monthly_start,
                        resolved_monthly_end,
                    ),
                    "customer_concentration_12mo": build_customer_concentration_query(
                        family,
                        resolved_customer_start,
                        resolved_customer_end,
                    ),
                },
            }
        )
        payload_template.append(
            build_payload_template_row(
                sku,
                family,
                row_numbers,
                resolved_monthly_start,
                resolved_monthly_end,
                resolved_customer_start,
                resolved_customer_end,
            )
        )
        for row_number in row_numbers:
            row_map[row_number].append(sku)

    return {
        "generated_at": utc_now(),
        "source_path": str(target.resolve()),
        "windows": {
            "monthly_sales_24mo": {
                "start_date": resolved_monthly_start,
                "end_date": resolved_monthly_end,
            },
            "customer_concentration_12mo": {
                "start_date": resolved_customer_start,
                "end_date": resolved_customer_end,
            },
        },
        "selection_guidance": {
            "monthly_sales_24mo": "Persist raw month-by-channel rows under monthly_sales. Each row should include month, sales_channel_id, revenue, units, and distinct_customers.",
            "customer_concentration_12mo": "Summarize the customer-grain query into the customer_concentration.amazon and customer_concentration.shopify blocks. Do the Pareto and repeat-rate math outside SQL before filling the payload.",
        },
        "sku_count": len(items),
        "row_count": len(row_map),
        "items": items,
        "payload_template": payload_template,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build batch Postgres MCP family-metrics queries for a workbook or session."
    )
    parser.add_argument(
        "target",
        nargs="?",
        default=str(DEFAULT_WORKBOOK),
        help="Workbook path or research session root.",
    )
    parser.add_argument(
        "--sheet",
        default=SHEET_NAME,
        help=f"Worksheet name when reading directly from a workbook (default: {SHEET_NAME}).",
    )
    parser.add_argument("--monthly-start-date", default=None, help="Override monthly-sales window start date (YYYY-MM-DD).")
    parser.add_argument("--monthly-end-date", default=None, help="Override monthly-sales window end date (YYYY-MM-DD).")
    parser.add_argument("--customer-start-date", default=None, help="Override customer-concentration window start date (YYYY-MM-DD).")
    parser.add_argument("--customer-end-date", default=None, help="Override customer-concentration window end date (YYYY-MM-DD).")
    parser.add_argument(
        "--output-prefix",
        default=None,
        help="Optional output prefix path without extension. Defaults next to the source target.",
    )
    args = parser.parse_args()

    target = Path(args.target).resolve()
    bundle = build_query_bundle(
        target=target,
        sheet_name=args.sheet,
        monthly_start_date=args.monthly_start_date,
        monthly_end_date=args.monthly_end_date,
        customer_start_date=args.customer_start_date,
        customer_end_date=args.customer_end_date,
    )

    output_prefix = Path(args.output_prefix).resolve() if args.output_prefix else output_stem_for(target)
    queries_path = output_prefix.with_name(output_prefix.name + "_postgres_queries.json")
    payload_path = output_prefix.with_name(output_prefix.name + "_payload_template.json")

    queries_path.write_text(
        json.dumps(
            {
                "generated_at": bundle["generated_at"],
                "source_path": bundle["source_path"],
                "windows": bundle["windows"],
                "selection_guidance": bundle["selection_guidance"],
                "sku_count": bundle["sku_count"],
                "row_count": bundle["row_count"],
                "items": bundle["items"],
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    payload_path.write_text(json.dumps(bundle["payload_template"], indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "source_path": bundle["source_path"],
                "sku_count": bundle["sku_count"],
                "row_count": bundle["row_count"],
                "windows": bundle["windows"],
                "queries_file": str(queries_path),
                "payload_template_file": str(payload_path),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
