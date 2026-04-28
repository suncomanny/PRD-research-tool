"""
Build a batch of Postgres MCP reference-baseline queries for a workbook/session.

Usage:
  python tools/reference_postgres_batch.py "C:\\path\\to\\filled_workbook.xlsx"
  python tools/reference_postgres_batch.py "C:\\path\\to\\output\\research_sessions\\session_name"
"""

from __future__ import annotations

import argparse
import json
from collections import defaultdict
from datetime import datetime, UTC
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from sku_lookup import build_mcp_queries, default_sales_window, strip_pack_suffix
from template_parser import DEFAULT_WORKBOOK, SHEET_NAME, normalize_header



def utc_now() -> str:
    """Return an ISO timestamp for artifact generation."""
    return datetime.now(UTC).replace(microsecond=0).isoformat()


def output_stem_for(target: Path) -> Path:
    """Choose a reasonable default output location for the generated files."""
    if target.is_dir():
        return target / "reference_postgres"
    return target.with_suffix("")


def build_payload_template_row(sku: str, family: str, row_numbers: list[int]) -> dict[str, Any]:
    """Return the merge-ready payload skeleton accepted by --postgres-json."""
    return {
        "sku": sku,
        "family": family,
        "row_numbers": row_numbers,
        "title": None,
        "listing_price": None,
        "shopify_revenue": None,
        "shopify_units": None,
        "amazon_revenue": None,
        "amazon_units": None,
    }


def collect_reference_rows_from_session(session_dir: Path) -> list[dict[str, Any]]:
    """Read unique reference SKUs from session packet files."""
    packets_dir = session_dir / "packets"
    references: dict[str, dict[str, Any]] = {}

    for packet_file in sorted(packets_dir.glob("row_*_packet.json")):
        packet = json.loads(packet_file.read_text(encoding="utf-8"))
        identity = packet.get("identity") or {}
        sku = identity.get("sunco_reference_sku")
        row_number = packet.get("row_number")
        if not sku or row_number is None:
            continue
        normalized_sku = str(sku).strip().upper()
        entry = references.setdefault(
            normalized_sku,
            {
                "sku": normalized_sku,
                "family": strip_pack_suffix(normalized_sku),
                "row_numbers": [],
            },
        )
        if row_number not in entry["row_numbers"]:
            entry["row_numbers"].append(int(row_number))

    return sorted(references.values(), key=lambda item: item["row_numbers"][0])


def collect_reference_rows_from_workbook(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Read unique reference SKUs directly from a workbook."""
    workbook = load_workbook(workbook_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

    ws = workbook[sheet_name]
    header_index: dict[int, str] = {}
    current_section = None

    for col in range(1, ws.max_column + 1):
        section_key = normalize_header(ws.cell(row=1, column=col).value)
        if section_key:
            current_section = section_key
        header_key = normalize_header(ws.cell(row=2, column=col).value)
        if header_key:
            header_index[col] = header_key if current_section == "ideation_identity" or current_section == "identity" else header_key

    sku_columns = [idx for idx, key in header_index.items() if key == "sunco_reference_sku"]
    if not sku_columns:
        raise ValueError("Could not find a 'Sunco Reference SKU' column in the workbook.")
    sku_column = sku_columns[0]

    references: dict[str, dict[str, Any]] = {}
    for row in range(3, ws.max_row + 1):
        raw_value = ws.cell(row=row, column=sku_column).value
        if raw_value is None:
            continue
        normalized_sku = str(raw_value).strip().upper()
        if not normalized_sku:
            continue
        entry = references.setdefault(
            normalized_sku,
            {
                "sku": normalized_sku,
                "family": strip_pack_suffix(normalized_sku),
                "row_numbers": [],
            },
        )
        entry["row_numbers"].append(row)

    return sorted(references.values(), key=lambda item: item["row_numbers"][0])


def collect_reference_rows(target: Path, sheet_name: str) -> list[dict[str, Any]]:
    """Collect unique reference SKUs from either a session dir or workbook."""
    if target.is_dir():
        packets_dir = target / "packets"
        if packets_dir.exists():
            return collect_reference_rows_from_session(target)
        raise ValueError(f"Directory '{target}' is not a research session root.")
    return collect_reference_rows_from_workbook(target, sheet_name)


def build_query_bundle(
    target: Path,
    sheet_name: str,
    start_date: str | None,
    end_date: str | None,
) -> dict[str, Any]:
    """Build the batch query payload plus a merge-ready template."""
    references = collect_reference_rows(target, sheet_name)
    resolved_start, resolved_end = (
        (start_date, end_date)
        if start_date and end_date
        else default_sales_window()
    )

    items = []
    payload_template = []
    row_map: dict[int, list[str]] = defaultdict(list)
    for ref in references:
        sku = ref["sku"]
        family = ref["family"]
        row_numbers = sorted(ref["row_numbers"])
        queries = build_mcp_queries(sku, start_date=resolved_start, end_date=resolved_end)
        items.append(
            {
                "sku": sku,
                "family": family,
                "row_numbers": row_numbers,
                "queries": queries,
            }
        )
        payload_template.append(build_payload_template_row(sku, family, row_numbers))
        for row_number in row_numbers:
            row_map[row_number].append(sku)

    return {
        "generated_at": utc_now(),
        "source_path": str(target.resolve()),
        "sales_window": {
            "start_date": resolved_start,
            "end_date": resolved_end,
        },
        "selection_guidance": {
            "title": "Use the first non-empty title row returned. Queries are ranked exact SKU first, then family-level matches.",
            "listing_price": "Prefer the first non-null list_price row returned. Queries are ranked by exact/family SKU match, then Shopify channel 12585 before Amazon channel 11929, then shorter listing SKU.",
            "sales_12mo": "Map sales_channel_id 12585 to Shopify revenue/units and 11929 to Amazon revenue/units.",
        },
        "sku_count": len(items),
        "row_count": len(row_map),
        "items": items,
        "payload_template": payload_template,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build batch Postgres MCP reference-baseline queries for a workbook or session."
    )
    parser.add_argument(
        "target",
        nargs="?",
        default=str(DEFAULT_WORKBOOK),
        help="Workbook path or research session root.",
    )
    parser.add_argument(
        "--sheet",
        default=SHEET_NAME,
        help=f"Worksheet name when reading directly from a workbook (default: {SHEET_NAME}).",
    )
    parser.add_argument(
        "--start-date",
        default=None,
        help="Override sales window start date (YYYY-MM-DD).",
    )
    parser.add_argument(
        "--end-date",
        default=None,
        help="Override sales window end date (YYYY-MM-DD).",
    )
    parser.add_argument(
        "--output-prefix",
        default=None,
        help="Optional output prefix path without extension. Defaults next to the source target.",
    )
    args = parser.parse_args()

    target = Path(args.target).resolve()
    bundle = build_query_bundle(
        target=target,
        sheet_name=args.sheet,
        start_date=args.start_date,
        end_date=args.end_date,
    )

    output_prefix = Path(args.output_prefix).resolve() if args.output_prefix else output_stem_for(target)
    queries_path = output_prefix.with_name(output_prefix.name + "_queries.json")
    payload_path = output_prefix.with_name(output_prefix.name + "_payload_template.json")

    queries_path.write_text(
        json.dumps(
            {
                "generated_at": bundle["generated_at"],
                "source_path": bundle["source_path"],
                "sales_window": bundle["sales_window"],
                "selection_guidance": bundle["selection_guidance"],
                "sku_count": bundle["sku_count"],
                "row_count": bundle["row_count"],
                "items": bundle["items"],
            },
            indent=2,
        ),
        encoding="utf-8",
    )
    payload_path.write_text(json.dumps(bundle["payload_template"], indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "source_path": bundle["source_path"],
                "sku_count": bundle["sku_count"],
                "row_count": bundle["row_count"],
                "sales_window": bundle["sales_window"],
                "queries_file": str(queries_path),
                "payload_template_file": str(payload_path),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
