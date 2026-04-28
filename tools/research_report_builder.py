"""
Step 6A: Build Excel research reports from completed analysis artifacts.

Usage:
  python tools/research_report_builder.py "C:\\path\\to\\research_session"
  python tools/research_report_builder.py "C:\\path\\to\\research_session" --rows 3,4,5
  python tools/research_report_builder.py "C:\\path\\to\\research_session" --combined
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

from research_session_manager import artifact_path_for, packet_path_for, read_json, update_session


HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E78")
SUBHEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")
ACCENT_FILL = PatternFill(fill_type="solid", fgColor="EAF2F8")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=12)
SUBHEADER_FONT = Font(bold=True)
TITLE_FONT = Font(bold=True, size=15)
HYPERLINK_FONT = Font(color="0563C1", underline="single")
WRAP_ALIGNMENT = Alignment(vertical="top", wrap_text=True)
DEFAULT_COLUMN_WIDTHS = {
    "A": 22,
    "B": 28,
    "C": 18,
    "D": 18,
    "E": 18,
    "F": 18,
    "G": 18,
    "H": 24,
    "I": 18,
    "J": 36,
    "K": 14,
}
ALTERNATE_ROW_FILL = PatternFill(fill_type="solid", fgColor="F7FBFF")
AMAZON_CHANNELS = {"amazon"}
BM_DIRECT_CHANNELS = {"home_depot", "walmart", "lowes", "brand_site", "stackline_seed"}
CHANNEL_DISPLAY_NAMES = {
    "amazon": "Amazon",
    "home_depot": "Home Depot",
    "walmart": "Walmart",
    "lowes": "Lowe's",
    "brand_site": "Brand Site",
    "stackline_seed": "Stackline",
}


def normalize_text(value: Any) -> str:
    """Render values as readable worksheet strings."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "Yes" if value else "No"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, list):
        return ", ".join([normalize_text(item) for item in value if normalize_text(item)])
    return str(value)


def as_dict(value: Any) -> dict[str, Any]:
    """Coerce optional mappings into dicts."""
    if isinstance(value, dict):
        return value
    return {}


def as_list(value: Any) -> list[Any]:
    """Coerce optional sequences into lists."""
    if isinstance(value, list):
        return value
    return []


def set_default_layout(ws) -> None:
    """Apply shared column widths and wrapping."""
    ws.freeze_panes = "A4"
    for column, width in DEFAULT_COLUMN_WIDTHS.items():
        ws.column_dimensions[column].width = width


def safe_sheet_title(base: str, used: set[str]) -> str:
    """Create a workbook-safe, unique worksheet title."""
    text = (base or "Report").replace("/", " ").replace("\\", " ").replace(":", " ").strip()
    text = text[:31] or "Report"
    candidate = text
    suffix = 2
    while candidate in used:
        trimmed = text[: max(0, 31 - len(f" ({suffix})"))]
        candidate = f"{trimmed} ({suffix})"
        suffix += 1
    used.add(candidate)
    return candidate


def section_header(ws, row: int, title: str, end_column: int = 10) -> int:
    """Write a section header row and return the next row index."""
    ws.cell(row=row, column=1, value=title)
    ws.cell(row=row, column=1).fill = HEADER_FILL
    ws.cell(row=row, column=1).font = HEADER_FONT
    ws.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=end_column)
    return row + 1


def key_value_rows(ws, row: int, pairs: list[tuple[str, Any]], columns: int = 2) -> int:
    """Write compact key/value pairs across the sheet."""
    index = 0
    while index < len(pairs):
        for block in range(columns):
            if index >= len(pairs):
                break
            label, value = pairs[index]
            base_col = (block * 2) + 1
            ws.cell(row=row, column=base_col, value=label)
            ws.cell(row=row, column=base_col).font = SUBHEADER_FONT
            ws.cell(row=row, column=base_col).fill = SUBHEADER_FILL
            ws.cell(row=row, column=base_col + 1, value=normalize_text(value))
            ws.cell(row=row, column=base_col + 1).alignment = WRAP_ALIGNMENT
            index += 1
        row += 1
    return row


def merged_text_row(ws, row: int, label: str, value: Any) -> int:
    """Write one labeled merged text row."""
    ws.cell(row=row, column=1, value=label)
    ws.cell(row=row, column=1).font = SUBHEADER_FONT
    ws.cell(row=row, column=1).fill = SUBHEADER_FILL
    ws.cell(row=row, column=2, value=normalize_text(value))
    ws.cell(row=row, column=2).alignment = WRAP_ALIGNMENT
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=10)
    return row + 1


def write_list_section(ws, row: int, title: str, values: list[str]) -> int:
    """Write a simple bulleted list section."""
    row = section_header(ws, row, title)
    if not values:
        ws.cell(row=row, column=1, value="No items.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        return row + 2
    for value in values:
        ws.cell(row=row, column=1, value=f"- {value}")
        ws.cell(row=row, column=1).alignment = WRAP_ALIGNMENT
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1
    return row + 1


def write_table(ws, row: int, title: str, headers: list[str], rows: list[list[Any]]) -> int:
    """Write a basic table and return the next row index."""
    row = section_header(ws, row, title, end_column=max(10, len(headers)))
    for col_index, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_index, value=header)
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        cell.alignment = WRAP_ALIGNMENT
    row += 1

    if not rows:
        ws.cell(row=row, column=1, value="No rows.")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(headers))
        return row + 2

    for row_offset, values in enumerate(rows):
        for col_index, value in enumerate(values, start=1):
            cell = ws.cell(row=row, column=col_index)
            apply_table_cell(cell, value)
            cell.alignment = WRAP_ALIGNMENT
            if row_offset % 2 == 1:
                cell.fill = ALTERNATE_ROW_FILL
        row += 1
    return row + 1


def apply_table_cell(cell, value: Any) -> None:
    """Write a table cell, optionally attaching a hyperlink."""
    hyperlink = None
    text_value = value
    if isinstance(value, dict):
        text_value = value.get("value")
        hyperlink = value.get("hyperlink")

    cell.value = normalize_text(text_value)
    if hyperlink:
        cell.hyperlink = hyperlink
        cell.font = HYPERLINK_FONT


def optional_text(value: Any) -> str | None:
    """Return a stripped string or None."""
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def looks_like_amazon_asin(value: str | None) -> bool:
    """Return whether a value looks like an Amazon ASIN."""
    return bool(value and re.fullmatch(r"[A-Z0-9]{10}", value.upper()))


def source_channel_label(item: dict[str, Any]) -> str:
    """Render a human-friendly source channel label."""
    channel = (optional_text(item.get("source_channel")) or "").lower()
    return CHANNEL_DISPLAY_NAMES.get(channel, channel.replace("_", " ").title() or "Unknown")


def listing_identifier_from_url(url: str | None, source_channel: str | None) -> str | None:
    """Best-effort identifier fallback derived from the listing URL."""
    if not url:
        return None

    channel = (source_channel or "").lower()
    patterns = []
    if channel == "amazon":
        patterns = [(r"/dp/([A-Z0-9]{10})(?:[/?]|$)", "ASIN {match}")]
    elif channel == "home_depot":
        patterns = [(r"/p/(?:[^/]+/)?(\d+)(?:[/?]|$)", "Item {match}")]
    elif channel == "walmart":
        patterns = [(r"/ip/(?:[^/]+/)?(\d+)(?:[/?]|$)", "Item {match}")]
    elif channel == "lowes":
        patterns = [(r"/pd/[^/]+/(\d+)(?:[/?]|$)", "Item {match}")]

    for pattern, template in patterns:
        match = re.search(pattern, url, flags=re.IGNORECASE)
        if match:
            return template.format(match=match.group(1))
    return None


def listing_identifier(item: dict[str, Any]) -> str:
    """Return the best channel-aware identifier label for one competitor listing."""
    source_channel = (optional_text(item.get("source_channel")) or "").lower()
    sku = optional_text(item.get("sku"))
    model_number = optional_text(item.get("model_number"))
    url = optional_text(item.get("url"))

    if source_channel == "amazon":
        asin = sku if looks_like_amazon_asin(sku) else None
        if not asin and looks_like_amazon_asin(model_number):
            asin = model_number
        if not asin and url:
            derived = listing_identifier_from_url(url, source_channel)
            if derived and derived.startswith("ASIN "):
                asin = derived.replace("ASIN ", "", 1)
        if asin:
            return f"ASIN {asin}"
        if model_number:
            return f"Model {model_number}"

    if model_number:
        return f"Model {model_number}"

    if sku:
        if source_channel == "brand_site":
            return f"Part {sku}"
        if source_channel in {"home_depot", "walmart", "lowes"}:
            return f"Item {sku}"
        return f"SKU {sku}"

    derived = listing_identifier_from_url(url, source_channel)
    return derived or ""


def listing_link_cell(item: dict[str, Any]) -> dict[str, str] | str:
    """Return a hyperlink cell payload for a competitor source listing."""
    url = optional_text(item.get("url"))
    if not url or url.startswith("stackline://"):
        return ""
    return {
        "value": source_channel_label(item),
        "hyperlink": url,
    }


def sort_candidates(items: list[dict[str, Any]]) -> list[dict[str, Any]]:
    """Sort candidate records by confidence, then by price."""
    def sort_key(item: dict[str, Any]) -> tuple[float, float, str]:
        confidence = item.get("match_confidence")
        if not isinstance(confidence, (int, float)):
            confidence = 0
        price = item.get("price")
        if not isinstance(price, (int, float)):
            price = 0
        title = normalize_text(item.get("product_title"))
        return (-confidence, price, title)

    return sorted(items, key=sort_key)


def candidate_rows(
    normalized_items: list[dict[str, Any]],
    channels: set[str],
    limit: int = 10,
) -> list[list[Any]]:
    """Build a compact competitor table filtered by source channel."""
    rows = []
    for item in sort_candidates(normalized_items):
        if normalize_text(item.get("source_channel")) not in channels:
            continue
        rows.append(
            [
                item.get("brand"),
                item.get("product_title"),
                listing_identifier(item),
                source_channel_label(item),
                item.get("price"),
                item.get("wattage"),
                item.get("lumens"),
                item.get("cct"),
                item.get("cri"),
                listing_link_cell(item),
                item.get("match_confidence"),
            ]
        )
        if len(rows) >= limit:
            break
    return rows


def top_brand_rows(brands: list[dict[str, Any]]) -> list[list[Any]]:
    """Format summarized brand rows."""
    rows = []
    for brand in brands:
        rows.append(
            [
                brand.get("brand"),
                brand.get("candidate_count"),
                normalize_text(brand.get("source_channels")),
                brand.get("median_unit_price"),
            ]
        )
    return rows


def coverage_rows(entries: list[dict[str, Any]]) -> list[list[Any]]:
    """Format feature/certification coverage rows."""
    rows = []
    for entry in entries:
        rows.append([entry.get("label"), entry.get("matched_count"), entry.get("coverage_pct")])
    return rows


def benchmark_rows(pricing: dict[str, Any]) -> list[list[Any]]:
    """Format multi-metric pricing benchmark rows."""
    metrics = [
        ("Raw Price", as_dict(pricing.get("price_benchmarks"))),
        ("Unit Price", as_dict(pricing.get("unit_price_benchmarks"))),
        ("Unit Price / Watt", as_dict(pricing.get("unit_price_per_watt_benchmarks"))),
        ("Unit Price / Lumen", as_dict(pricing.get("unit_price_per_lumen_benchmarks"))),
    ]
    rows = []
    for label, metric in metrics:
        rows.append(
            [
                label,
                metric.get("sample_size"),
                metric.get("min"),
                metric.get("p25"),
                metric.get("median"),
                metric.get("mean"),
                metric.get("p75"),
                metric.get("max"),
            ]
        )
    return rows


def pricing_position_rows(pricing: dict[str, Any]) -> list[list[Any]]:
    """Format pricing positioning rows."""
    suggested = as_dict(pricing.get("suggested_msrp_range"))
    target_position = as_dict(pricing.get("target_price_position"))
    return [
        ["Target MSRP", pricing.get("target_msrp")],
        ["Evaluated Price", target_position.get("evaluated_value")],
        ["Evaluated Price Source", target_position.get("evaluated_value_source")],
        ["Target Price Percentile", target_position.get("percentile")],
        ["Target Price Bucket", target_position.get("bucket")],
        ["Target vs Market Median %", target_position.get("vs_median_pct")],
        ["Observed Unit Price Floor (P25)", suggested.get("observed_unit_price_floor")],
        ["Observed Unit Price Ceiling (P75)", suggested.get("observed_unit_price_ceiling")],
        ["Recommended Floor", suggested.get("recommended_floor")],
        ["Recommended Ceiling", suggested.get("recommended_ceiling")],
        ["Minimum Margin-Safe MSRP", suggested.get("minimum_margin_safe_price")],
        ["Suggested Positioning", suggested.get("positioning")],
        ["Margin Conflict", suggested.get("margin_conflict")],
    ]


def margin_rows(pricing: dict[str, Any]) -> list[list[Any]]:
    """Format channel-specific margin guidance rows."""
    rows = []
    for channel in ["shopify", "amazon"]:
        entry = as_dict(as_dict(pricing.get("margin_targets")).get(channel))
        if not entry:
            continue
        rows.append(
            [
                channel.title(),
                entry.get("target_margin_pct"),
                entry.get("minimum_viable_msrp"),
                entry.get("vs_target_msrp_pct"),
                entry.get("vs_market_median_pct"),
            ]
        )
    return rows


def value_position_rows(pricing: dict[str, Any]) -> list[list[Any]]:
    """Format value-ranking rows for unit price, price per watt, and price per lumen."""
    rows = []
    metrics = [
        ("Unit Price", as_dict(pricing.get("target_price_position"))),
        ("Unit Price / Watt", as_dict(pricing.get("target_price_per_watt_position"))),
        ("Unit Price / Lumen", as_dict(pricing.get("target_price_per_lumen_position"))),
    ]
    for label, metric in metrics:
        if not metric:
            continue
        rows.append(
            [
                label,
                metric.get("evaluated_value"),
                metric.get("percentile"),
                metric.get("bucket"),
                metric.get("vs_median_pct"),
            ]
        )
    return rows


def channel_comparison_rows(performance: dict[str, Any]) -> list[list[Any]]:
    """Format Stackline channel comparison rows for Section A."""
    comparison = as_dict(performance.get("channel_comparison"))
    channels = as_dict(comparison.get("channels"))
    rows = []
    for channel_name, channel in channels.items():
        channel = as_dict(channel)
        rows.append(
            [
                channel_name,
                channel.get("retail_sales"),
                channel.get("units_sold"),
                channel.get("avg_retail_price"),
                channel.get("retail_sales_growth_pct"),
                channel.get("sunco_sales_share_pct"),
            ]
        )
    return rows


def spec_action_rows(spec_coverage: dict[str, Any]) -> list[list[Any]]:
    """Format actionable feature/certification coverage rows."""
    rows = []
    for entry in as_list(spec_coverage.get("feature_coverage")):
        rows.append(
            [
                "Feature",
                entry.get("label"),
                entry.get("signal"),
                entry.get("evidence_strength"),
                entry.get("coverage_pct"),
                entry.get("matched_count"),
                entry.get("recommended_action"),
            ]
        )
    for entry in as_list(spec_coverage.get("certification_coverage")):
        rows.append(
            [
                "Certification",
                entry.get("label"),
                entry.get("signal"),
                entry.get("evidence_strength"),
                entry.get("coverage_pct"),
                entry.get("matched_count"),
                entry.get("recommended_action"),
            ]
        )
    return rows


def numeric_guidance_rows(spec_coverage: dict[str, Any]) -> list[list[Any]]:
    """Format numeric target-positioning rows."""
    rows = []
    for entry in as_list(spec_coverage.get("numeric_guidance")):
        rows.append(
            [
                entry.get("label"),
                entry.get("target_value"),
                entry.get("median"),
                entry.get("p75"),
                entry.get("target_percentile"),
                entry.get("recommended_action"),
            ]
        )
    return rows


def gate_readiness_snapshot_rows(gate_readiness: dict[str, Any]) -> list[list[Any]]:
    """Format gate/channel readiness rows for the report."""
    rows = []
    for snapshot in as_list(gate_readiness.get("snapshots")):
        evidence = as_dict(snapshot.get("evidence_confidence"))
        rows.append(
            [
                snapshot.get("channel"),
                snapshot.get("gate"),
                snapshot.get("family_state"),
                snapshot.get("weighted_score"),
                evidence.get("score"),
                evidence.get("label"),
                f"{evidence.get('implemented_questions')}/{evidence.get('methodology_active_questions')}",
            ]
        )
    return rows


def gate_readiness_pillar_rows(gate_readiness: dict[str, Any]) -> list[list[Any]]:
    """Format pillar-level rollups from the primary G2 channel snapshot."""
    primary_channel = normalize_text(gate_readiness.get("primary_channel"))
    for snapshot in as_list(gate_readiness.get("snapshots")):
        if snapshot.get("channel") == primary_channel and snapshot.get("gate") == "G2":
            rows = []
            for pillar in as_list(snapshot.get("pillar_scores")):
                rows.append(
                    [
                        pillar.get("label"),
                        pillar.get("base_weight"),
                        pillar.get("effective_weight"),
                        pillar.get("average_score"),
                        f"{pillar.get('scored_question_count')}/{pillar.get('question_count')}",
                        pillar.get("status"),
                    ]
                )
            return rows
    return []


def vendor_request_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    """Format vendor optimization requests."""
    rows = []
    for item in items:
        rows.append(
            [
                item.get("priority"),
                item.get("linked_metric"),
                item.get("request"),
                item.get("reason"),
            ]
        )
    return rows


def optimization_driver_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    """Format category-aware decision drivers for the report."""
    rows = []
    for item in items:
        rows.append(
            [
                item.get("tier"),
                item.get("label"),
                item.get("driver_type"),
                item.get("signal"),
                item.get("reason"),
            ]
        )
    return rows


def low_signal_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    """Format lower-signal attributes that should be validated before over-weighting."""
    rows = []
    for item in items:
        rows.append(
            [
                item.get("label"),
                item.get("driver_type"),
                item.get("signal"),
                item.get("reason"),
            ]
        )
    return rows


def optimization_modifier_rows(items: list[dict[str, Any]]) -> list[list[Any]]:
    """Format active optimization modifiers for the report."""
    rows = []
    for item in items:
        rows.append(
            [
                item.get("label"),
                ", ".join(as_list(item.get("matched_keywords"))),
                item.get("match_score"),
            ]
        )
    return rows


def optimization_scorecard_rows(scorecard: dict[str, Any]) -> list[list[Any]]:
    """Format optimization score components for the report."""
    rows = []
    for item in as_list(scorecard.get("components")):
        rows.append(
            [
                item.get("component"),
                item.get("score"),
                item.get("weight"),
                item.get("reason"),
            ]
        )
    return rows


def report_filename(row_number: int) -> str:
    """Return the stable report filename for a row."""
    return f"row_{row_number:03d}_research_report.xlsx"


def prd_prefill_pairs(packet: dict[str, Any], analysis: dict[str, Any]) -> list[tuple[str, Any]]:
    """Build the PRD-oriented prefill block from packet targets."""
    identity = as_dict(packet.get("identity"))
    target_profile = as_dict(packet.get("target_profile"))
    electrical = as_dict(target_profile.get("electrical"))
    physical = as_dict(target_profile.get("physical"))
    business = as_dict(target_profile.get("business_case"))
    reference = as_dict(packet.get("reference_baseline"))

    return [
        ("Ideation Name", identity.get("ideation_name")),
        ("Reference Image URL", reference.get("image_url")),
        ("Voltage", electrical.get("voltage")),
        ("Wattage Primary", electrical.get("wattage_primary")),
        ("Wattage Max", electrical.get("wattage_max")),
        ("Selectable Wattage", electrical.get("selectable_wattage")),
        ("CCT Primary", electrical.get("cct_primary")),
        ("CCT Max", electrical.get("cct_max")),
        ("Selectable CCT", electrical.get("selectable_cct")),
        ("CRI", electrical.get("cri")),
        ("Lumens Target", electrical.get("lumens_target")),
        ("Dimmable", electrical.get("dimmable")),
        ("Dimming Type", electrical.get("dimming_type")),
        ("Size / Form Factor", physical.get("size_form_factor")),
        ("Mounting Type", physical.get("mounting_type")),
        ("Material", physical.get("material")),
        ("Finish / Color", physical.get("finish_color")),
        ("IP Rating", physical.get("ip_rating")),
        ("Moisture Rating", physical.get("moisture_rating")),
        ("Target MSRP", as_dict(analysis.get("pricing_analysis")).get("target_msrp")),
        ("Target Vendor Cost", as_dict(analysis.get("pricing_analysis")).get("target_vendor_cost")),
        ("Certifications", business.get("certifications")),
        ("Lifetime Hours", business.get("lifetime_hours")),
        ("Warranty", business.get("warranty")),
    ]


def render_row_sheet(
    ws,
    row_number: int,
    packet: dict[str, Any],
    analysis: dict[str, Any],
    normalized: dict[str, Any],
) -> None:
    """Render one ideation sheet into an existing worksheet."""
    set_default_layout(ws)

    identity = as_dict(packet.get("identity"))
    reference = as_dict(packet.get("reference_baseline"))
    performance = as_dict(analysis.get("performance_estimation"))
    pricing = as_dict(analysis.get("pricing_analysis"))
    summary = as_dict(analysis.get("summary"))
    spec_coverage = as_dict(analysis.get("spec_coverage"))
    reference_anchor = as_dict(analysis.get("reference_anchor_context"))
    gate_readiness = as_dict(analysis.get("gate_readiness"))
    vendor_requests = as_list(analysis.get("highest_impact_vendor_requests"))
    ideation_optimization = as_dict(analysis.get("ideation_optimization"))
    optimization_scorecard = as_dict(ideation_optimization.get("optimization_scorecard"))
    optimization_modifiers = as_list(ideation_optimization.get("active_modifiers"))

    ws.cell(row=1, column=1, value=identity.get("ideation_name"))
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
    ws.cell(row=2, column=1, value=f"Row {row_number} Research Report")
    ws.cell(row=2, column=1).fill = ACCENT_FILL
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=10)

    row = 4
    row = section_header(ws, row, "Section A - Ideation + Reference Anchor Context")
    row = key_value_rows(
        ws,
        row,
        [
            ("Category Owner", identity.get("category_owner")),
            ("Category", identity.get("category")),
            ("Subcategory", identity.get("subcategory")),
            ("Strategy", identity.get("strategy")),
            ("Reference Anchor SKU", identity.get("sunco_reference_sku")),
            ("Launch Outlook", performance.get("launch_outlook")),
            ("Confidence", performance.get("confidence")),
            ("Anchor Data Quality", reference_anchor.get("data_quality")),
            ("Anchor Title", reference.get("title")),
            ("Anchor Title Source", reference.get("title_source")),
            ("Anchor Listing Price", reference.get("listing_price")),
            ("Listing Price Source", reference.get("listing_price_source")),
            ("Listing Price Note", reference.get("listing_price_note")),
            ("Anchor Shopify Revenue 12mo", reference.get("shopify_revenue_12mo")),
            ("Anchor Shopify Units 12mo", reference.get("shopify_units_12mo")),
            ("Shopify Data Source", reference.get("shopify_data_source")),
            ("Anchor Amazon Revenue 12mo", reference.get("amazon_revenue_12mo")),
            ("Anchor Amazon Units 12mo", reference.get("amazon_units_12mo")),
            ("Amazon Data Source", reference.get("amazon_data_source")),
            ("Anchor Data Source", reference.get("reference_data_source")),
            ("Anchor Sales Period", reference.get("sales_period_label")),
        ],
    )
    row = merged_text_row(ws, row, "Reference Anchor Role", reference_anchor.get("primary_use"))
    row = merged_text_row(ws, row, "Reference Anchor Secondary Use", reference_anchor.get("secondary_use"))
    row = merged_text_row(ws, row, "Reference Anchor Caution", reference_anchor.get("caution"))
    row = merged_text_row(ws, row, "Reference Anchor Guardrail", reference_anchor.get("do_not_overweight"))
    row = merged_text_row(ws, row, "Reference Anchor Image URL", reference.get("image_url"))
    row = merged_text_row(
        ws,
        row,
        "Performance Rationale",
        " | ".join(as_list(performance.get("rationale"))),
    )
    row = write_table(
        ws,
        row,
        "Stackline Channel Comparison",
        ["Channel", "Retail Sales", "Units Sold", "Avg Price", "Sales Growth %", "Sunco Share %"],
        channel_comparison_rows(performance),
    )
    row = merged_text_row(ws, row, "Gate Readiness Summary", gate_readiness.get("summary"))
    row = write_table(
        ws,
        row,
        "Gate Readiness by Channel / Gate",
        ["Channel", "Gate", "Family State", "Score", "Evidence Score", "Evidence Label", "Implemented Questions"],
        gate_readiness_snapshot_rows(gate_readiness),
    )
    row = write_table(
        ws,
        row,
        "Primary G2 Pillar Rollup",
        ["Pillar", "Base Weight", "Effective Weight", "Avg Score", "Scored Questions", "Status"],
        gate_readiness_pillar_rows(gate_readiness),
    )

    row = write_table(
        ws,
        row,
        "Section B - Amazon Competitors",
        ["Brand", "Product", "Identifier", "Channel", "Price", "Wattage", "Lumens", "CCT", "CRI", "Source Link", "Confidence"],
        candidate_rows(as_list(normalized.get("items")), AMAZON_CHANNELS, limit=10),
    )

    row = write_table(
        ws,
        row,
        "Section C - Brick-and-Mortar / Direct Competitors",
        ["Brand", "Product", "Identifier", "Channel", "Price", "Wattage", "Lumens", "CCT", "CRI", "Source Link", "Confidence"],
        candidate_rows(as_list(normalized.get("items")), BM_DIRECT_CHANNELS, limit=12),
    )

    row = write_table(
        ws,
        row,
        "Section D - Pricing Position",
        ["Metric", "Value"],
        pricing_position_rows(pricing),
    )
    row = write_table(
        ws,
        row,
        "Pricing Benchmarks",
        ["Metric", "Samples", "Min", "P25", "Median", "Mean", "P75", "Max"],
        benchmark_rows(pricing),
    )
    row = write_table(
        ws,
        row,
        "Margin Targets",
        ["Channel", "Target Margin %", "Min MSRP", "Vs Target MSRP %", "Vs Market Median %"],
        margin_rows(pricing),
    )
    row = write_table(
        ws,
        row,
        "Value Ranking",
        ["Metric", "Target", "Percentile", "Bucket", "Vs Median %"],
        value_position_rows(pricing),
    )
    row = write_table(
        ws,
        row,
        "Top Brands",
        ["Brand", "Candidate Count", "Channels", "Median Unit Price"],
        top_brand_rows(as_list(summary.get("top_brands"))),
    )

    row = write_table(
        ws,
        row,
        "Section E - Feature / Certification Signals",
        ["Type", "Label", "Signal", "Evidence", "Coverage %", "Matched", "Recommendation"],
        spec_action_rows(spec_coverage),
    )
    row = write_table(
        ws,
        row,
        "Numeric Target Positioning",
        ["Metric", "Target", "Median", "P75", "Percentile", "Recommendation"],
        numeric_guidance_rows(spec_coverage),
    )
    row = merged_text_row(ws, row, "Category Optimization Summary", ideation_optimization.get("summary"))
    row = key_value_rows(
        ws,
        row,
        [
            ("Optimization Profile", ideation_optimization.get("profile_label")),
            ("Profile Match Basis", ", ".join(as_list(ideation_optimization.get("matched_taxonomy"))) or ", ".join(as_list(ideation_optimization.get("matched_keywords")))),
            ("Active Modifiers", ", ".join(item.get("label") for item in optimization_modifiers if item.get("label"))),
            ("Optimization Score", optimization_scorecard.get("score")),
            ("Optimization Label", optimization_scorecard.get("label")),
            ("Optimization Confidence", optimization_scorecard.get("confidence")),
        ],
    )
    row = write_table(
        ws,
        row,
        "Active Variant Modifiers",
        ["Modifier", "Matched Keywords", "Match Score"],
        optimization_modifier_rows(optimization_modifiers),
    )
    row = write_table(
        ws,
        row,
        "Optimization Scorecard",
        ["Component", "Score", "Weight", "Reason"],
        optimization_scorecard_rows(optimization_scorecard),
    )
    row = write_table(
        ws,
        row,
        "Primary Decision Drivers",
        ["Tier", "Driver", "Type", "Signal", "Why It Matters"],
        optimization_driver_rows(as_list(ideation_optimization.get("primary_decision_drivers"))),
    )
    row = write_table(
        ws,
        row,
        "Secondary Decision Drivers",
        ["Tier", "Driver", "Type", "Signal", "Why It Matters"],
        optimization_driver_rows(as_list(ideation_optimization.get("secondary_decision_drivers"))),
    )
    row = write_table(
        ws,
        row,
        "Validate Before Over-Weighting",
        ["Driver", "Type", "Signal", "Reason"],
        low_signal_rows(as_list(ideation_optimization.get("low_signal_attributes"))),
    )
    row = write_table(
        ws,
        row,
        "Highest-Impact Vendor Requests",
        ["Priority", "Area", "Request", "Reason"],
        vendor_request_rows(vendor_requests),
    )
    row = write_list_section(ws, row, "Recommendations", as_list(analysis.get("recommendations")))
    row = write_list_section(ws, row, "Notes", as_list(analysis.get("notes")))

    row = section_header(ws, row, "Section F - PRD Generator Pre-Fill")
    row = key_value_rows(ws, row, prd_prefill_pairs(packet, analysis))


def completed_row_payloads(
    session_dir: Path,
    rows: list[int] | None = None,
) -> list[tuple[int, dict[str, Any], dict[str, Any], dict[str, Any]]]:
    """Load completed packet/analysis/normalized payloads for selected rows."""
    manifest = read_json(session_dir / "manifest.json")
    target_rows = set(rows or [row["row_number"] for row in manifest.get("rows", [])])
    payloads = []

    for row in manifest.get("rows", []):
        row_number = row["row_number"]
        if row_number not in target_rows:
            continue
        if row.get("stages", {}).get("analyzed") != "complete":
            continue
        packet = read_json(packet_path_for(session_dir, row_number))
        analysis = read_json(artifact_path_for(session_dir, row_number, "analyzed"))
        normalized = read_json(artifact_path_for(session_dir, row_number, "normalized"))
        payloads.append((row_number, packet, analysis, normalized))

    return payloads


def build_report(session_dir: Path, row_number: int) -> Path | None:
    """Build one Excel report for a completed analyzed row."""
    payloads = completed_row_payloads(session_dir, rows=[row_number])
    if not payloads:
        return None

    _, packet, analysis, normalized = payloads[0]
    wb = Workbook()
    ws = wb.active
    ws.title = safe_sheet_title(normalize_text(as_dict(packet.get("identity")).get("ideation_name")), set())
    render_row_sheet(ws, row_number, packet, analysis, normalized)

    report_path = artifact_path_for(session_dir, row_number, "reported")
    report_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(report_path)
    return report_path


def build_summary_sheet(ws, payloads: list[tuple[int, dict[str, Any], dict[str, Any], dict[str, Any]]]) -> None:
    """Build a summary sheet for the combined workbook."""
    set_default_layout(ws)
    ws.title = "Summary"
    ws.cell(row=1, column=1, value="Completed Research Rows")
    ws.cell(row=1, column=1).font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=9)
    key_value_rows(
        ws,
        3,
        [
            ("Completed Row Count", len(payloads)),
            ("Workbook Scope", "Completed analyzed rows only"),
        ],
        columns=1,
    )
    write_table(
        ws,
        6,
        "Completed Rows",
        ["Row", "Ideation", "Category Owner", "Category", "Outlook", "Confidence", "Amazon G2", "Amazon Evidence", "Report File"],
        [
            [
                row_number,
                as_dict(packet.get("identity")).get("ideation_name"),
                as_dict(packet.get("identity")).get("category_owner"),
                as_dict(packet.get("identity")).get("category"),
                as_dict(analysis.get("performance_estimation")).get("launch_outlook"),
                as_dict(analysis.get("performance_estimation")).get("confidence"),
                next(
                    (
                        snapshot.get("weighted_score")
                        for snapshot in as_list(as_dict(analysis.get("gate_readiness")).get("snapshots"))
                        if snapshot.get("channel") == "amazon" and snapshot.get("gate") == "G2"
                    ),
                    None,
                ),
                next(
                    (
                        as_dict(snapshot.get("evidence_confidence")).get("label")
                        for snapshot in as_list(as_dict(analysis.get("gate_readiness")).get("snapshots"))
                        if snapshot.get("channel") == "amazon" and snapshot.get("gate") == "G2"
                    ),
                    None,
                ),
                report_filename(row_number),
            ]
            for row_number, packet, analysis, _ in payloads
        ],
    )


def build_combined_workbook(
    session_dir: Path,
    rows: list[int] | None = None,
    output_path: str | None = None,
) -> Path | None:
    """Build one workbook with a summary sheet plus one sheet per completed row."""
    payloads = completed_row_payloads(session_dir, rows=rows)
    if not payloads:
        return None

    wb = Workbook()
    summary_ws = wb.active
    build_summary_sheet(summary_ws, payloads)

    used_titles = {summary_ws.title}
    for row_number, packet, analysis, normalized in payloads:
        title = safe_sheet_title(normalize_text(as_dict(packet.get("identity")).get("ideation_name")), used_titles)
        ws = wb.create_sheet(title=title)
        render_row_sheet(ws, row_number, packet, analysis, normalized)

    if output_path:
        destination = Path(output_path).resolve()
    else:
        destination = session_dir / "reports" / f"{session_dir.name}_completed_rows.xlsx"
    destination.parent.mkdir(parents=True, exist_ok=True)
    wb.save(destination)
    return destination


def parse_rows_argument(value: str | None) -> list[int] | None:
    """Parse an optional comma-separated row list."""
    if not value:
        return None
    rows = []
    for part in value.split(","):
        part = part.strip()
        if not part:
            continue
        rows.append(int(part))
    return rows or None


def build_reports(
    session_root: str,
    rows: list[int] | None = None,
    combined: bool = False,
    output_path: str | None = None,
) -> dict[str, Any]:
    """Build report artifacts for selected session rows."""
    session_dir = Path(session_root).resolve()

    if combined:
        combined_path = build_combined_workbook(session_dir, rows=rows, output_path=output_path)
        update_result = update_session(str(session_dir))
        return {
            "session_root": str(session_dir),
            "rows_requested": sorted(rows or []),
            "combined_workbook": str(combined_path) if combined_path else None,
            "manifest_summary": update_result["summary"],
        }

    manifest = read_json(session_dir / "manifest.json")
    target_rows = set(rows or [row["row_number"] for row in manifest.get("rows", [])])

    written_rows = []
    skipped_rows = []
    report_files = []

    for row in manifest.get("rows", []):
        row_number = row["row_number"]
        if row_number not in target_rows:
            continue
        report_path = build_report(session_dir, row_number)
        if report_path is None:
            skipped_rows.append(row_number)
            continue
        written_rows.append(row_number)
        report_files.append(str(report_path))

    update_result = update_session(str(session_dir))
    return {
        "session_root": str(session_dir),
        "rows_requested": sorted(target_rows),
        "rows_written": written_rows,
        "rows_skipped": skipped_rows,
        "report_files": report_files,
        "manifest_summary": update_result["summary"],
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Build Excel report artifacts for completed analyzed rows."
    )
    parser.add_argument("session_root", help="Path to an initialized research session.")
    parser.add_argument(
        "--rows",
        default=None,
        help="Optional comma-separated row numbers to report.",
    )
    parser.add_argument(
        "--combined",
        action="store_true",
        help="Build one combined workbook with one sheet per completed row.",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Optional explicit output path for the combined workbook.",
    )
    args = parser.parse_args()

    result = build_reports(
        args.session_root,
        rows=parse_rows_argument(args.rows),
        combined=args.combined,
        output_path=args.output,
    )
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()
