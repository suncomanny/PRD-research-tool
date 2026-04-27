"""
Step 3: Parse the PRD research template and attach Reference SKU context.

Usage:
  python tools/template_parser.py templates/PRD_Research_Template.xlsx
  python tools/template_parser.py templates/PRD_Research_Template.xlsx --postgres-json sales.json
  python tools/template_parser.py templates/PRD_Research_Template.xlsx --include-queries
  python tools/template_parser.py templates/PRD_Research_Template_Test.xlsx --include-stackline-raw
"""

from __future__ import annotations

import argparse
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from stackline_analyzer import (
    STACKLINE_DIR,
    analyze_stackline_channels_for_subcategory,
)
from sku_lookup import build_mcp_queries, lookup_from_csv, merge_postgres_data


TOOLS_DIR = Path(__file__).resolve().parent
REPO_ROOT = TOOLS_DIR.parent
DEFAULT_WORKBOOK = REPO_ROOT / "templates" / "PRD_Research_Template.xlsx"
SHEET_NAME = "Ideations"
REQUIRED_IDENTITY_FIELDS = {
    "category",
    "subcategory",
    "ideation_name",
    "sunco_reference_sku",
    "strategy",
}
BOOLEAN_FIELDS = {
    "selectable_wattage",
    "selectable_cct",
    "dimmable",
    "emergency_battery",
    "motion_sensor",
    "daylight_sensor_auto_dimming",
    "smart_connected",
    "linkable",
    "stackline_data",
}
LIST_FIELDS = {
    "known_competitors",
    "priority_channels",
    "certifications",
}
SECTION_KEY_OVERRIDES = {
    "core_electrical_specs": "electrical_specs",
    "features_and_requirements": "features_requirements",
}


def stackline_expected(research_guidance: dict[str, Any]) -> bool:
    """Treat Stackline as the default unless the workbook explicitly disables it."""
    return research_guidance.get("stackline_data") is not False


def normalize_header(value: Any) -> str | None:
    """Normalize a worksheet header into a stable snake_case key."""
    if value is None:
        return None

    text = str(value).replace("*", "").strip().lower()
    if not text:
        return None

    text = text.replace("&", "and")
    text = text.replace("%", "pct")
    text = text.replace("/", " ")
    text = re.sub(r"[()?'-]+", " ", text)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_") or None


def normalize_value(value: Any) -> Any:
    """Trim worksheet values and coerce empty strings to None."""
    if value is None:
        return None
    if isinstance(value, str):
        text = value.strip()
        return text or None
    return value


def coerce_field_value(field_key: str, value: Any) -> Any:
    """Apply lightweight type coercion for common template field patterns."""
    normalized = normalize_value(value)
    if normalized is None:
        return None

    if field_key in BOOLEAN_FIELDS and isinstance(normalized, str):
        lowered = normalized.lower()
        if lowered == "yes":
            return True
        if lowered == "no":
            return False

    return normalized


def split_list_value(value: Any) -> list[str] | None:
    """Split comma- or semicolon-delimited user text into a cleaned list."""
    if not isinstance(value, str):
        return None

    parts = [
        part.strip()
        for part in re.split(r",|;", value)
        if part and part.strip()
    ]
    return parts or None


def build_column_specs(worksheet) -> list[dict[str, Any]]:
    """Read sections from row 1 and column headers from row 2."""
    specs = []
    current_section = None

    for column_index in range(1, worksheet.max_column + 1):
        section_value = normalize_header(worksheet.cell(row=1, column=column_index).value)
        if section_value:
            current_section = SECTION_KEY_OVERRIDES.get(section_value, section_value)

        header_value = worksheet.cell(row=2, column=column_index).value
        header_key = normalize_header(header_value)
        if not header_key:
            continue

        specs.append(
            {
                "column_index": column_index,
                "header": str(header_value).strip(),
                "key": header_key,
                "section": current_section or "unsectioned",
            }
        )

    return specs


def load_sku_payloads(path: str | None) -> dict[str, dict[str, Any]]:
    """Load per-SKU payloads from a dict or list JSON artifact."""
    if not path:
        return {}

    data = json.loads(Path(path).read_text(encoding="utf-8"))
    payloads: dict[str, dict[str, Any]] = {}

    if isinstance(data, dict):
        if "sku" in data and isinstance(data.get("sku"), str):
            sku = data["sku"].strip().upper()
            payloads[sku] = {k: v for k, v in data.items() if k != "sku"}
            return payloads

        for sku, payload in data.items():
            if isinstance(payload, dict):
                payloads[str(sku).strip().upper()] = payload
        return payloads

    if isinstance(data, list):
        for item in data:
            if not isinstance(item, dict):
                continue
            sku = item.get("sku")
            if not sku:
                continue
            payloads[str(sku).strip().upper()] = {
                k: v for k, v in item.items() if k != "sku"
            }
        return payloads

    raise ValueError("Unsupported payload JSON format. Use a dict or list of dicts.")


def load_postgres_payloads(path: str | None) -> dict[str, dict[str, Any]]:
    """Load per-SKU Postgres enrichment payloads from JSON."""
    return load_sku_payloads(path)


def load_family_metrics_payloads(path: str | None) -> dict[str, dict[str, Any]]:
    """Load per-SKU family metrics enrichment payloads from JSON."""
    return load_sku_payloads(path)


def merge_family_metrics_data(
    reference_context: dict[str, Any] | None,
    family_metrics_payload: dict[str, Any] | None,
) -> dict[str, Any]:
    """Merge monthly/customer family metrics into the reference context."""
    merged = dict(reference_context or {})
    if not family_metrics_payload:
        return merged

    for key in (
        "monthly_sales",
        "customer_concentration",
        "sales_trend",
        "gmc_visibility",
        "family_metrics_source",
        "family_metrics_period_label",
        "family_metrics_notes",
    ):
        value = family_metrics_payload.get(key)
        if value not in (None, "", [], {}):
            merged[key] = value

    merged["family_metrics_enrichment_provided"] = True
    return merged


def build_stackline_context(
    identity: dict[str, Any],
    research_guidance: dict[str, Any],
    reference_sku: str | None,
    stackline_folder: Path,
    stackline_brand: str,
    include_stackline_raw: bool,
) -> tuple[dict[str, Any] | None, list[str]]:
    """Attach Stackline market context for ideation performance estimation."""
    if not stackline_expected(research_guidance):
        return {
            "enabled": False,
            "expected": False,
            "matched": False,
            "mode": "web_only",
            "fallback_mode": "web_collection_only",
            "warnings": ["Stackline explicitly disabled for this row."],
        }, []

    subcategory = identity.get("subcategory")
    if not subcategory:
        return {
            "enabled": True,
            "expected": True,
            "matched": False,
            "mode": "web_fallback",
            "fallback_mode": "web_collection_only",
            "subcategory": None,
            "warnings": ["Stackline expected but subcategory is blank."],
        }, ["Stackline expected but subcategory is blank."]

    try:
        stackline_batch = analyze_stackline_channels_for_subcategory(
            subcategory=str(subcategory),
            reference_sku=reference_sku,
            folder=stackline_folder,
            brand_name=stackline_brand,
        )
    except FileNotFoundError as exc:
        return {
            "enabled": True,
            "expected": True,
            "matched": False,
            "mode": "web_fallback",
            "fallback_mode": "web_collection_only",
            "subcategory": subcategory,
            "warnings": [str(exc)],
        }, [str(exc)]
    except Exception as exc:
        return {
            "enabled": True,
            "expected": True,
            "matched": False,
            "mode": "web_fallback",
            "fallback_mode": "web_collection_only",
            "subcategory": subcategory,
            "warnings": [f"Stackline analysis failed: {exc}"],
        }, [f"Stackline analysis failed: {exc}"]

    analysis = stackline_batch.get("primary_analysis") or {}
    channels = stackline_batch.get("channels") or {}
    channel_performance = {
        channel: payload.get("performance_estimation_context")
        for channel, payload in channels.items()
        if payload.get("performance_estimation_context")
    }
    channel_files = {
        channel: (payload.get("analysis") or {}).get("files")
        for channel, payload in channels.items()
        if (payload.get("analysis") or {}).get("files")
    }

    context = {
        "enabled": True,
        "expected": True,
        "matched": True,
        "mode": "stackline_first",
        "fallback_mode": "targeted_web_enrichment",
        "subcategory": subcategory,
        "primary_channel": stackline_batch.get("primary_channel"),
        "segment_name": analysis.get("segment_name"),
        "matched_bundle": analysis.get("matched_bundle"),
        "files": analysis.get("files"),
        "channel_files": channel_files,
        "performance_estimation_context": analysis.get("performance_estimation_context"),
        "channel_performance_estimation_contexts": channel_performance,
        "channel_comparison": stackline_batch.get("channel_comparison"),
        "channels": {
            channel: {
                "matched_bundle": (payload.get("analysis") or {}).get("matched_bundle"),
                "files": (payload.get("analysis") or {}).get("files"),
                "performance_estimation_context": payload.get("performance_estimation_context"),
            }
            for channel, payload in channels.items()
        },
        "warnings": stackline_batch.get("warnings", []),
    }
    if include_stackline_raw:
        context["raw_analysis"] = analysis
        context["channel_raw_analysis"] = {
            channel: payload.get("analysis")
            for channel, payload in channels.items()
        }

    return context, []


def parse_template(
    workbook_path: str,
    postgres_payloads: dict[str, dict[str, Any]] | None = None,
    family_metrics_payloads: dict[str, dict[str, Any]] | None = None,
    include_queries: bool = False,
    include_stackline_raw: bool = False,
    start_date: str | None = None,
    end_date: str | None = None,
    stackline_folder: str | None = None,
    stackline_brand: str = "Sunco Lighting",
    sheet_name: str = SHEET_NAME,
) -> dict[str, Any]:
    """Parse the ideations sheet into enriched ideation objects."""
    payloads = postgres_payloads or {}
    family_metrics = family_metrics_payloads or {}
    resolved_stackline_folder = Path(stackline_folder) if stackline_folder else STACKLINE_DIR
    workbook = load_workbook(workbook_path, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in workbook.")

    worksheet = workbook[sheet_name]
    column_specs = build_column_specs(worksheet)
    ideations = []
    warnings = []

    for row_number in range(3, worksheet.max_row + 1):
        row_sections: dict[str, dict[str, Any]] = {}
        row_has_data = False

        for spec in column_specs:
            raw_value = worksheet.cell(row=row_number, column=spec["column_index"]).value
            value = coerce_field_value(spec["key"], raw_value)
            if value is not None:
                row_has_data = True

            section_bucket = row_sections.setdefault(spec["section"], {})
            section_bucket[spec["key"]] = value

            if spec["key"] in LIST_FIELDS and value is not None:
                section_bucket[f"{spec['key']}_list"] = split_list_value(value)

        if not row_has_data:
            continue

        identity = row_sections.setdefault("identity", {})
        reference_sku = identity.get("sunco_reference_sku")
        issues = []

        missing_required = [
            field.replace("_", " ")
            for field in REQUIRED_IDENTITY_FIELDS
            if not identity.get(field)
        ]
        if missing_required:
            issues.append(
                "Missing required fields: " + ", ".join(sorted(missing_required))
            )

        reference_context = None
        normalized_sku = None
        if reference_sku:
            normalized_sku = str(reference_sku).strip().upper()
            reference_context = lookup_from_csv(normalized_sku)

            postgres_payload = payloads.get(normalized_sku)
            if postgres_payload:
                reference_context = merge_postgres_data(reference_context, postgres_payload)

            family_metrics_payload = family_metrics.get(normalized_sku)
            if family_metrics_payload:
                reference_context = merge_family_metrics_data(
                    reference_context,
                    family_metrics_payload,
                )

            if include_queries:
                reference_context["postgres_queries"] = build_mcp_queries(
                    normalized_sku,
                    start_date=start_date,
                    end_date=end_date,
                )

            reference_context["postgres_enrichment_provided"] = bool(postgres_payload)
            reference_context["family_metrics_enrichment_provided"] = bool(family_metrics_payload)

            if not reference_context.get("found"):
                issues.append(
                    f"Reference SKU '{normalized_sku}' was not found in the metadata CSV."
                )

        stackline_context, stackline_issues = build_stackline_context(
            identity=identity,
            research_guidance=row_sections.get("research_guidance", {}),
            reference_sku=normalized_sku,
            stackline_folder=resolved_stackline_folder,
            stackline_brand=stackline_brand,
            include_stackline_raw=include_stackline_raw,
        )
        issues.extend(stackline_issues)

        ideations.append(
            {
                "row_number": row_number,
                "identity": row_sections.get("identity", {}),
                "electrical_specs": row_sections.get("electrical_specs", {}),
                "physical_mechanical": row_sections.get("physical_mechanical", {}),
                "features_requirements": row_sections.get("features_requirements", {}),
                "business_targets": row_sections.get("business_targets", {}),
                "research_guidance": row_sections.get("research_guidance", {}),
                "reference_context": reference_context,
                "stackline_context": stackline_context,
                "issues": issues,
            }
        )

        if issues:
            warnings.append({"row_number": row_number, "issues": issues})

    return {
        "workbook_path": str(Path(workbook_path).resolve()),
        "sheet_name": sheet_name,
        "ideation_count": len(ideations),
        "ideations": ideations,
        "warnings": warnings,
    }


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Parse the PRD research template and attach Reference SKU data."
    )
    parser.add_argument(
        "workbook",
        nargs="?",
        default=str(DEFAULT_WORKBOOK),
        help="Path to the filled PRD research workbook.",
    )
    parser.add_argument(
        "--sheet",
        default=SHEET_NAME,
        help=f"Worksheet name to parse (default: {SHEET_NAME}).",
    )
    parser.add_argument(
        "--postgres-json",
        default=None,
        help="JSON file with per-SKU Postgres enrichment payloads.",
    )
    parser.add_argument(
        "--family-metrics-json",
        default=None,
        help="JSON file with per-SKU monthly sales and customer concentration payloads.",
    )
    parser.add_argument(
        "--include-queries",
        action="store_true",
        help="Include MCP query templates for each parsed Reference SKU.",
    )
    parser.add_argument(
        "--include-stackline-raw",
        action="store_true",
        help="Include the full raw Stackline analysis alongside the compact estimation context.",
    )
    parser.add_argument(
        "--start-date",
        default=None,
        help="Override MCP sales query start date (YYYY-MM-DD).",
    )
    parser.add_argument(
        "--end-date",
        default=None,
        help="Override MCP sales query end date (YYYY-MM-DD).",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Optional path to write the JSON output.",
    )
    parser.add_argument(
        "--stackline-folder",
        default=str(STACKLINE_DIR),
        help="Folder containing Stackline summary / traffic exports.",
    )
    parser.add_argument(
        "--stackline-brand",
        default="Sunco Lighting",
        help="Brand name to treat as the internal brand focus in Stackline analysis.",
    )
    args = parser.parse_args()

    postgres_payloads = load_postgres_payloads(args.postgres_json)
    family_metrics_payloads = load_family_metrics_payloads(args.family_metrics_json)
    parsed = parse_template(
        workbook_path=args.workbook,
        postgres_payloads=postgres_payloads,
        family_metrics_payloads=family_metrics_payloads,
        include_queries=args.include_queries,
        include_stackline_raw=args.include_stackline_raw,
        start_date=args.start_date,
        end_date=args.end_date,
        stackline_folder=args.stackline_folder,
        stackline_brand=args.stackline_brand,
        sheet_name=args.sheet,
    )

    output = json.dumps(parsed, indent=2)
    if args.output:
        output_path = Path(args.output)
        output_path.write_text(output + "\n", encoding="utf-8")
    print(output)


if __name__ == "__main__":
    main()
