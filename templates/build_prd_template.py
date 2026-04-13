"""
Build PRD Research Template (.xlsx) for Sunco Lighting.
Generates a 2-sheet workbook: Instructions + Ideations data entry.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── Colors ──────────────────────────────────────────────────────────────
DARK_BLUE = "1F4E79"
LIGHT_BLUE = "D6E4F0"
WHITE = "FFFFFF"
ALT_ROW = "F2F2F2"
RED_REQUIRED = "C00000"

dark_blue_fill = PatternFill(start_color=DARK_BLUE, end_color=DARK_BLUE, fill_type="solid")
light_blue_fill = PatternFill(start_color=LIGHT_BLUE, end_color=LIGHT_BLUE, fill_type="solid")
alt_fill = PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")
white_fill = PatternFill(start_color=WHITE, end_color=WHITE, fill_type="solid")

white_bold_12 = Font(name="Calibri", size=12, bold=True, color=WHITE)
header_font = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
header_req_font = Font(name="Calibri", size=10, bold=True, color=RED_REQUIRED)
body_font = Font(name="Calibri", size=10)
instr_title = Font(name="Calibri", size=16, bold=True, color=DARK_BLUE)
instr_heading = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
instr_body = Font(name="Calibri", size=11)
instr_bold = Font(name="Calibri", size=11, bold=True)

thin_border = Border(
    bottom=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
)

wrap_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = openpyxl.Workbook()

# ════════════════════════════════════════════════════════════════════════
# SHEET 1: Instructions
# ════════════════════════════════════════════════════════════════════════
ws_instr = wb.active
ws_instr.title = "Instructions"
ws_instr.sheet_properties.tabColor = DARK_BLUE

# Column widths
ws_instr.column_dimensions["A"].width = 4
ws_instr.column_dimensions["B"].width = 90

# Title row
ws_instr.merge_cells("B2:B2")
c = ws_instr["B2"]
c.value = "PRD Research Template \u2014 Instructions"
c.font = instr_title
ws_instr.row_dimensions[2].height = 35

# Intro
row = 4
intro_lines = [
    ("This template captures product ideation data for the PRD Research Tool.", instr_body),
    ("Fill in one row per ideation product on the \u201cIdeations\u201d sheet.", instr_body),
    ("The tool will look up additional data from the Reference SKU you provide.", instr_body),
    ("", instr_body),
    ("REQUIRED FIELDS (marked with * on the Ideations sheet):", instr_bold),
    ("  \u2022 Category, Subcategory, Ideation Name, Sunco Reference SKU, Strategy", instr_body),
    ("", instr_body),
    ("SECTION OVERVIEW:", instr_heading),
    ("", instr_body),
    ("IDENTITY (Columns A\u2013F)", instr_bold),
    ("  Core product identification: what category it belongs to, its working name,", instr_body),
    ("  the reference SKU to look up existing data, and the strategic intent.", instr_body),
    ("", instr_body),
    ("CORE ELECTRICAL SPECS (Columns G\u2013U)", instr_bold),
    ("  Electrical performance targets: voltage, wattage, color temperature, CRI,", instr_body),
    ("  lumens output, dimming, driver type. Mark selectable fields Yes/No.", instr_body),
    ("", instr_body),
    ("PHYSICAL / MECHANICAL (Columns V\u2013AC)", instr_bold),
    ("  Form factor, mounting, materials, finish, IP/moisture ratings,", instr_body),
    ("  indoor/outdoor suitability, and operating temperature range.", instr_body),
    ("", instr_body),
    ("FEATURES & REQUIREMENTS (Columns AD\u2013AO)", instr_bold),
    ("  Special features: emergency battery, motion sensor, daylight sensor,", instr_body),
    ("  smart connectivity, linkability, bulb type, and beam angle.", instr_body),
    ("", instr_body),
    ("BUSINESS TARGETS (Columns AP\u2013AV)", instr_bold),
    ("  Pricing targets (MSRP, margin, vendor cost), certifications,", instr_body),
    ("  rated lifetime, and warranty period.", instr_body),
    ("", instr_body),
    ("RESEARCH GUIDANCE (Columns AW\u2013AY)", instr_bold),
    ("  Known competitors to benchmark, priority sales channels, and any", instr_body),
    ("  additional notes or context for the research tool.", instr_body),
    ("", instr_body),
    ("TIPS:", instr_heading),
    ("  \u2022 Fill in what you know \u2014 leave unknown fields blank for auto-lookup.", instr_body),
    ("  \u2022 Use the dropdown menus for standardized fields.", instr_body),
    ("  \u2022 The Reference SKU drives the data lookup \u2014 make sure it\u2019s accurate.", instr_body),
    ("  \u2022 Row 3 contains an example entry you can overwrite.", instr_body),
]

for text, font in intro_lines:
    cell = ws_instr.cell(row=row, column=2, value=text)
    cell.font = font
    cell.alignment = Alignment(vertical="center", wrap_text=True)
    row += 1

# Header bar at top
ws_instr.merge_cells("A1:B1")
for col in range(1, 3):
    c = ws_instr.cell(row=1, column=col)
    c.fill = dark_blue_fill
ws_instr.row_dimensions[1].height = 8

# ════════════════════════════════════════════════════════════════════════
# SHEET 2: Ideations
# ════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("Ideations")
ws.sheet_properties.tabColor = DARK_BLUE

# ── Section definitions ─────────────────────────────────────────────────
sections = [
    ("IDENTITY", "A", "F"),
    ("CORE ELECTRICAL SPECS", "G", "U"),
    ("PHYSICAL / MECHANICAL", "V", "AC"),
    ("FEATURES & REQUIREMENTS", "AD", "AO"),
    ("BUSINESS TARGETS", "AP", "AV"),
    ("RESEARCH GUIDANCE", "AW", "AY"),
]

# ── Column definitions (col_letter, header_text, is_required) ────────
columns = [
    # IDENTITY
    ("A", "Category *", True),
    ("B", "Subcategory *", True),
    ("C", "Ideation Name *", True),
    ("D", "Sunco Reference SKU *", True),
    ("E", "Reference SKU Source", False),
    ("F", "Strategy *", True),
    # CORE ELECTRICAL SPECS
    ("G", "Voltage", False),
    ("H", "Wattage (Primary)", False),
    ("I", "Wattage (Max)", False),
    ("J", "Selectable Wattage?", False),
    ("K", "CCT (Primary)", False),
    ("L", "CCT (Max)", False),
    ("M", "Selectable CCT?", False),
    ("N", "CRI", False),
    ("O", "Lumens (Target)", False),
    ("P", "Efficiency (lm/W)", False),
    ("Q", "Power Factor", False),
    ("R", "Dimmable?", False),
    ("S", "Dimming Type", False),
    ("T", "Frequency", False),
    ("U", "Driver Type", False),
    # PHYSICAL / MECHANICAL
    ("V", "Size / Form Factor", False),
    ("W", "Mounting Type", False),
    ("X", "Material", False),
    ("Y", "Finish / Color", False),
    ("Z", "IP Rating", False),
    ("AA", "Moisture Rating", False),
    ("AB", "Indoor/Outdoor Use", False),
    ("AC", "Operating Temperature", False),
    # FEATURES & REQUIREMENTS
    ("AD", "Emergency / Battery?", False),
    ("AE", "Run Time", False),
    ("AF", "Charge Time", False),
    ("AG", "Switching Time", False),
    ("AH", "Motion Sensor?", False),
    ("AI", "Motion Duration", False),
    ("AJ", "Daylight Sensor / Auto-Dimming?", False),
    ("AK", "Smart / Connected?", False),
    ("AL", "Linkable?", False),
    ("AM", "Bulb Base Type", False),
    ("AN", "Bulb Shape", False),
    ("AO", "Beam Angle", False),
    # BUSINESS TARGETS
    ("AP", "Target MSRP", False),
    ("AQ", "Target Margin % (Shopify)", False),
    ("AR", "Target Margin % (Amazon)", False),
    ("AS", "Target Vendor Cost", False),
    ("AT", "Certifications", False),
    ("AU", "Lifetime Hours", False),
    ("AV", "Warranty", False),
    # RESEARCH GUIDANCE
    ("AW", "Known Competitors", False),
    ("AX", "Priority Channels", False),
    ("AY", "Research Notes", False),
]

# ── Column widths by section ────────────────────────────────────────────
width_map = {
    "IDENTITY": 18,
    "CORE ELECTRICAL SPECS": 15,
    "PHYSICAL / MECHANICAL": 16,
    "FEATURES & REQUIREMENTS": 15,
    "BUSINESS TARGETS": 16,
    "RESEARCH GUIDANCE": 22,
}

def col_index(letter):
    """Convert column letter to 1-based index."""
    result = 0
    for ch in letter:
        result = result * 26 + (ord(ch) - ord('A') + 1)
    return result

# Set column widths
for sec_name, start_let, end_let in sections:
    w = width_map[sec_name]
    s_idx = col_index(start_let)
    e_idx = col_index(end_let)
    for i in range(s_idx, e_idx + 1):
        ws.column_dimensions[get_column_letter(i)].width = w

# ── Row 1: Section headers (merged, dark blue) ─────────────────────────
ws.row_dimensions[1].height = 30
for sec_name, start_let, end_let in sections:
    ws.merge_cells(f"{start_let}1:{end_let}1")
    cell = ws[f"{start_let}1"]
    cell.value = sec_name
    cell.font = white_bold_12
    cell.fill = dark_blue_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    # Fill all cells in the merged range so the background shows
    s_idx = col_index(start_let)
    e_idx = col_index(end_let)
    for i in range(s_idx, e_idx + 1):
        c = ws.cell(row=1, column=i)
        c.fill = dark_blue_fill

# ── Row 2: Column headers ──────────────────────────────────────────────
ws.row_dimensions[2].height = 40
for col_let, header_text, required in columns:
    idx = col_index(col_let)
    cell = ws.cell(row=2, column=idx, value=header_text)
    cell.font = header_req_font if required else header_font
    cell.fill = light_blue_fill
    cell.alignment = wrap_align
    cell.border = thin_border

# ── Auto-filter on row 2 ───────────────────────────────────────────────
last_col_letter = "AY"
ws.auto_filter.ref = f"A2:{last_col_letter}2"

# ── Freeze panes: freeze rows 1-2 and columns A-C ─────────────────────
ws.freeze_panes = "D3"

# ── Alternating row fills (rows 3-102) ─────────────────────────────────
max_col = col_index("AY")
for row_num in range(3, 103):
    fill = alt_fill if row_num % 2 == 0 else white_fill
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row_num, column=c)
        cell.fill = fill
        cell.font = body_font
        cell.alignment = Alignment(vertical="center")
        cell.border = thin_border

# ── Data Validations ───────────────────────────────────────────────────
validations = {
    "A": "Accessory,Bulbs,Electrical,Grow Lights,High Bays + Low Bays,Indoor Commercial,Indoor Residential,Industrial,Outdoor Commercial,Outdoor Fixtures,Recessed,Tubes",
    "B": "A series,Accessory,Area Lights,B11 series,Bathroom Fans,BR series,Bulbs,Canless,Canopy,Cans,Ceiling Fixtures,Dimmers,Disk,Emergency,Flood Lights,G series,Lamps,Linears,MR series,Panels,PAR series,Pendants,Receptacles,Recessed,Residential,Residential Landscape,Retrofit,Retrofit Kits,S series,Safety,Sensors,Shop Lights,Smart Lights,ST series,Strip Lights,Tubes,UFO,Under Cabinet,Vanity,Vapor Tights,Wall Packs,Wall Sconces,Wire,Wraparounds",
    "F": "New Product,Revision,Cost Reduction,Vendor Transition",
    "G": "12V,24V,120V,120-277V,120-347V,200-480V,277-480V,Low Voltage,Other",
    "J": "Yes,No",
    "M": "Yes,No",
    "R": "Yes,No",
    "S": "0-10V,1-10V,Triac,DALI,PWM,App-controlled,Phase-cut,ELV,Other",
    "T": "50Hz,60Hz,50/60Hz",
    "U": "Driver,DOB (Driver On Board),N/A",
    "W": "Recessed,Surface,Flush Mount,Semi-flush,Pendant,Ceiling,Wall,Yoke,Pole,Ground Spike,Track,Chain/Hanging,Freestanding,Other",
    "X": "Aluminum,Steel,Die-cast Aluminum,Polycarbonate,ABS,Acrylic/PMMA,Glass,Brass,Cast Brass,Iron,Stainless Steel,Plastic,Other",
    "Y": "White,Black,Bronze,Antique Bronze,Brushed Nickel,Chrome,Clear,Frosted,Gun Metal,Stainless Steel,Matte,Silver,Other",
    "Z": "None/NA,IP20,IP40,IP44,IP54,IP65,IP66,IP67,IP69K",
    "AA": "Dry,Damp,Wet",
    "AB": "Indoor,Outdoor,Indoor/Outdoor",
    "AD": "Yes,No",
    "AH": "Yes,No",
    "AJ": "Yes,No",
    "AK": "Yes,No",
    "AL": "Yes,No",
    "AM": "E12,E26,E39,G5,G13,G24q,GU10,GU24,GU5.3,MR-16,Integrated LED,N/A,Other",
    "AV": "1-Year,2-Year,3-Year,5-Year,7-Year,10-Year,Lifetime",
    "AX": "Amazon,Home Depot,Walmart,Lowe's,Direct/Distributor,All",
}

for col_let, formula_list in validations.items():
    dv = DataValidation(
        type="list",
        formula1=f'"{formula_list}"',
        allow_blank=True,
        showDropDown=False,
    )
    dv.error = "Please select a value from the dropdown list."
    dv.errorTitle = "Invalid Entry"
    dv.prompt = "Select from list"
    dv.promptTitle = "Choose Value"
    dv.showInputMessage = True
    dv.showErrorMessage = True
    idx = col_index(col_let)
    letter = get_column_letter(idx)
    dv.add(f"{letter}3:{letter}102")
    ws.add_data_validation(dv)

# ── Example row (row 3) ────────────────────────────────────────────────
example = {
    "A": "Indoor Commercial",
    "B": "Panels",
    "C": "2x4 Panel Selectable 50W",
    "D": "PN24-2x4-30W-1PK",
    "E": "Existing Catalog",
    "F": "New Product",
    "G": "120-277V",
    "H": "50",
    "I": "60",
    "J": "Yes",
    "K": "4000K",
    "L": "5000K",
    "M": "Yes",
    "N": "80+",
    "O": "6000",
    "P": "120",
    "Q": ">0.9",
    "R": "Yes",
    "S": "0-10V",
    "T": "60Hz",
    "U": "Driver",
    "V": '2\'x4\' (24"x48")',
    "W": "Recessed",
    "X": "Aluminum",
    "Y": "White",
    "Z": "IP40",
    "AA": "Dry",
    "AB": "Indoor",
    "AC": "-20\u00b0C to 45\u00b0C",
    "AD": "No",
    "AE": "",
    "AF": "",
    "AG": "",
    "AH": "No",
    "AI": "",
    "AJ": "No",
    "AK": "No",
    "AL": "No",
    "AM": "Integrated LED",
    "AN": "",
    "AO": "120\u00b0",
    "AP": "$49.99",
    "AQ": "45%",
    "AR": "35%",
    "AS": "$18.00",
    "AT": "UL, DLC, FCC, Energy Star",
    "AU": "50,000",
    "AV": "7-Year",
    "AW": "Metalux, Lithonia, TCP",
    "AX": "Amazon",
    "AY": "High-output selectable wattage panel to replace PN24-2x4-30W line. Target DLC Premium for utility rebates.",
}

for col_let, value in example.items():
    idx = col_index(col_let)
    cell = ws.cell(row=3, column=idx, value=value)
    cell.font = body_font
    cell.alignment = Alignment(vertical="center")
    # Keep alternating fill (row 3 is odd = white)
    cell.fill = white_fill
    cell.border = thin_border

# ── Save ────────────────────────────────────────────────────────────────
import os
output_dir = r"C:\Users\Sunco\OneDrive - Sunco Lighting\Documents\Claude Workbook\PRD-research-tool\templates"
os.makedirs(output_dir, exist_ok=True)
output_path = os.path.join(output_dir, "PRD_Research_Template.xlsx")
wb.save(output_path)
print(f"SUCCESS: Template saved to {output_path}")
print(f"  - Sheet 1: Instructions")
print(f"  - Sheet 2: Ideations ({len(columns)} columns, example in row 3)")
print(f"  - Data validations: {len(validations)} dropdown columns")
print(f"  - Data rows prepared: 3-102 (100 rows)")
